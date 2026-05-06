import streamlit as st
import anthropic
import base64
import json
import io
from pathlib import Path

st.set_page_config(page_title="ADV — Extraction BC", page_icon="📋", layout="wide")

st.markdown("""
<style>
    .main { padding-top: 1.5rem; }
    div[data-testid="metric-container"] {
        background: #f8f8f6; border: 0.5px solid #e0dfd8;
        border-radius: 8px; padding: 12px 16px;
    }
</style>
""", unsafe_allow_html=True)

try:
    api_key = st.secrets["ANTHROPIC_API_KEY"]
except Exception:
    api_key = None

with st.sidebar:
    st.markdown("## ⚙️ À propos")
    st.markdown("MVP ADV — Extraction automatique de bons de commande.\n\n**Stack :** Streamlit · Claude API · openpyxl\n\n**Coût estimé :** ~0,01 € par BC analysé")
    st.markdown("---")
    st.caption("Outil en version bêta — usage démonstration uniquement.")

SYSTEM_PROMPT = """Tu es un assistant spécialisé dans l'analyse de bons de commande pour un ADV (Administration Des Ventes).
Extrait toutes les informations et réponds UNIQUEMENT en JSON valide, sans preamble ni backticks Markdown.

Champs PRIORITAIRES (ne jamais laisser null si présent dans le document) :
- destinataire : à qui ce BC est adressé (nom société/service destinataire — pas le client qui commande)
- client : nom de la société qui passe la commande
- site_livraison : nom du site ou établissement de livraison
- adresse_livraison : adresse postale complète de livraison
- date_livraison : date souhaitée de livraison
- heure_livraison : heure ou créneau horaire de livraison si mentionné
- numero_contrat : numéro de contrat cadre ou référence contrat
- lignes > reference_produit : référence produit fournisseur si mentionnée
- lignes > prix_unitaire : prix unitaire de chaque ligne
- lignes > designation : désignation exacte du produit

Structure JSON exacte :
{
  "destinataire": "Nom société/service à qui le BC est adressé",
  "client": "Nom complet de la société qui commande",
  "contact_client": "Nom du signataire ou contact côté client, ou null",
  "email_contact": "Email ou null",
  "telephone_contact": "Téléphone ou null",
  "numero_commande": "Numéro du BC client",
  "numero_contrat": "Numéro de contrat cadre, ou null",
  "date_commande": "JJ/MM/AAAA",
  "date_livraison": "JJ/MM/AAAA ou null",
  "heure_livraison": "HH:MM ou créneau ex: 8h-12h, ou null",
  "site_livraison": "Nom du site ou établissement, ou null",
  "adresse_livraison": "Adresse postale complète, ou null",
  "incoterms": "EXW / DAP / DDP / etc. ou null",
  "devise": "EUR",
  "lignes": [
    {
      "numero_ligne": 1,
      "reference_client": "Référence produit écrite par le client dans le BC",
      "reference_produit": "Référence produit fournisseur si mentionnée, ou null",
      "designation": "Désignation exacte du produit commandé",
      "quantite": 10,
      "unite": "pcs / kg / m / l / etc.",
      "prix_unitaire": 25.50,
      "total_ligne": 255.00,
      "date_livraison_ligne": "JJ/MM/AAAA si différente de la date globale, ou null"
    }
  ],
  "total_ht": 255.00,
  "tva_pct": 20,
  "total_ttc": 306.00,
  "conditions_paiement": "30 jours fin de mois ou null",
  "remarques": "Conditions particulières, urgences, ou null",
  "alertes": ["Liste des anomalies : mauvais destinataire, prix incohérent, date passée, quantité suspecte..."],
  "confiance": "haute / moyenne / basse"
}

RÈGLES :
- 'destinataire' est critique : si le nom ne correspond pas à un fournisseur habituel, signale-le dans 'alertes'.
- Si prix unitaire absent mais total et quantité présents, calcule le prix unitaire.
- Incohérences détectées → champ 'alertes'.
- Montants en nombres décimaux, jamais en chaînes.
- Information absente → null, ne jamais inventer."""


def extraire_bc(fichier_bytes: bytes, nom_fichier: str, cle_api: str) -> dict:
    client = anthropic.Anthropic(api_key=cle_api)
    ext = Path(nom_fichier).suffix.lower()
    if ext == ".pdf":
        bloc = {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": base64.standard_b64encode(fichier_bytes).decode("utf-8")}}
    elif ext in [".png", ".jpg", ".jpeg", ".webp"]:
        mt = f"image/{'jpeg' if ext in ['.jpg', '.jpeg'] else ext[1:]}"
        bloc = {"type": "image", "source": {"type": "base64", "media_type": mt, "data": base64.standard_b64encode(fichier_bytes).decode("utf-8")}}
    else:
        raise ValueError(f"Format non supporté : {ext}")

    reponse = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=2500,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": [bloc, {"type": "text", "text": "Analyse ce bon de commande et extrais toutes les informations en respectant les priorités."}]}],
    )
    texte = reponse.content[0].text.strip().replace("```json", "").replace("```", "").strip()
    return json.loads(texte)


def generer_excel(d: dict) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return b""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commande SAP"

    gris = PatternFill("solid", fgColor="F1EFE8")
    bleu = PatternFill("solid", fgColor="E6F1FB")
    rouge = PatternFill("solid", fgColor="FCEBEB")
    font_h = Font(bold=True, size=10)
    align_c = Alignment(horizontal="center", vertical="center")
    bord = Border(bottom=Side(style="thin", color="D3D1C7"), top=Side(style="thin", color="D3D1C7"))

    ws.merge_cells("A1:I1")
    ws["A1"] = "BON DE COMMANDE — SAISIE SAP"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = bleu
    ws["A1"].alignment = align_c

    row = 2
    alertes = d.get("alertes", [])
    if alertes:
        row += 1
        ws.merge_cells(f"A{row}:I{row}")
        ws.cell(row=row, column=1, value="ALERTES : " + " | ".join(alertes))
        ws.cell(row=row, column=1).font = Font(bold=True, size=10, color="A32D2D")
        for col in range(1, 10):
            ws.cell(row=row, column=col).fill = rouge

    row += 2
    infos = [
        ("Destinataire BC", d.get("destinataire", "")),
        ("Client", d.get("client", "")),
        ("Site livraison", d.get("site_livraison", "")),
        ("Adresse livraison", d.get("adresse_livraison", "")),
        ("Date livraison", d.get("date_livraison", "")),
        ("Heure livraison", d.get("heure_livraison", "")),
        ("N° contrat", d.get("numero_contrat", "")),
        ("N° commande client", d.get("numero_commande", "")),
        ("Date commande", d.get("date_commande", "")),
        ("Contact", d.get("contact_client", "")),
        ("Email", d.get("email_contact", "")),
        ("Téléphone", d.get("telephone_contact", "")),
        ("Incoterms", d.get("incoterms", "")),
        ("Conditions paiement", d.get("conditions_paiement", "")),
        ("Devise", d.get("devise", "EUR")),
    ]
    for label, val in infos:
        ws.cell(row=row, column=1, value=label).font = font_h
        ws.cell(row=row, column=1).fill = gris
        ws.cell(row=row, column=2, value=str(val) if val else "").font = Font(size=10)
        ws.merge_cells(f"B{row}:I{row}")
        row += 1

    row += 1
    headers = ["#", "Réf. client", "Réf. produit", "Désignation", "Qté", "Unité", "Prix U. (€)", "Total HT (€)", "Livraison ligne"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = font_h; c.fill = bleu; c.alignment = align_c; c.border = bord

    for ltr, w in zip("ABCDEFGHI", [4, 16, 16, 32, 7, 7, 13, 13, 16]):
        ws.column_dimensions[ltr].width = w

    row += 1
    for ligne in d.get("lignes", []):
        vals = [ligne.get("numero_ligne",""), ligne.get("reference_client",""), ligne.get("reference_produit",""),
                ligne.get("designation",""), ligne.get("quantite",""), ligne.get("unite",""),
                ligne.get("prix_unitaire",""), ligne.get("total_ligne",""), ligne.get("date_livraison_ligne","") or ""]
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=v)
        if row % 2 == 0:
            for col in range(1, 10):
                ws.cell(row=row, column=col).fill = gris
        row += 1

    row += 1
    ws.cell(row=row, column=7, value="Total HT").font = Font(bold=True)
    ws.cell(row=row, column=8, value=d.get("total_ht","")).font = Font(bold=True)
    row += 1
    tva = (d.get("total_ht") or 0) * (d.get("tva_pct", 20) / 100)
    ws.cell(row=row, column=7, value=f"TVA {d.get('tva_pct',20)}%").font = Font(bold=True)
    ws.cell(row=row, column=8, value=round(tva, 2)).font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=7, value="Total TTC").font = Font(bold=True)
    ws.cell(row=row, column=8, value=d.get("total_ttc","")).font = Font(bold=True)
    ws.cell(row=row, column=8).fill = bleu

    if d.get("remarques"):
        row += 2
        ws.cell(row=row, column=1, value="Remarques : " + d["remarques"]).font = Font(size=10, italic=True)
        ws.merge_cells(f"A{row}:I{row}")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def afficher_resultats(d: dict):
    symbole = "€" if d.get("devise", "EUR") == "EUR" else d.get("devise", "")

    def fmt(val):
        if val is None: return "—"
        try: return f"{float(val):,.2f} {symbole}".replace(",", " ")
        except: return str(val)

    # Alertes en rouge
    for a in d.get("alertes", []):
        st.error(f"⚠ {a}")

    # Vérification destinataire — affiché en vert si OK
    if d.get("destinataire"):
        st.success(f"📬 BC adressé à : **{d['destinataire']}**")

    # Métriques clés
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Client", d.get("client", "—"))
    c2.metric("N° commande", d.get("numero_commande", "—"))
    c3.metric("N° contrat", d.get("numero_contrat") or "—")
    c4.metric("Total HT", fmt(d.get("total_ht")))
    c5.metric("Lignes", len(d.get("lignes", [])))

    st.markdown("---")
    left, right = st.columns([1, 2])

    with left:
        st.markdown("##### Livraison")
        for label, val in [("Site", d.get("site_livraison")), ("Adresse", d.get("adresse_livraison")),
                           ("Date", d.get("date_livraison")), ("Heure / créneau", d.get("heure_livraison"))]:
            if val:
                st.markdown(f"**{label}** : {val}")

        st.markdown("##### Commande")
        for label, val in [("Date BC", d.get("date_commande")), ("Contact", d.get("contact_client")),
                           ("Email", d.get("email_contact")), ("Téléphone", d.get("telephone_contact")),
                           ("Incoterms", d.get("incoterms")), ("Paiement", d.get("conditions_paiement"))]:
            if val:
                st.markdown(f"**{label}** : {val}")

        icone = {"haute": "🟢", "moyenne": "🟡", "basse": "🔴"}.get(d.get("confiance", "moyenne"), "🟡")
        st.markdown(f"\n**Confiance** : {icone} {(d.get('confiance','moyenne')).capitalize()}")
        if d.get("remarques"):
            st.info(f"**Remarques**\n\n{d['remarques']}")

    with right:
        st.markdown("##### Lignes de commande")
        for ligne in d.get("lignes", []):
            ref_cli = ligne.get("reference_client", "")
            ref_pro = ligne.get("reference_produit", "")
            refs = ref_cli + (f" → {ref_pro}" if ref_pro and ref_pro != ref_cli else "")
            c1, c2, c3, c4 = st.columns([3, 1, 1, 2])
            c1.markdown(f"**{refs}**  \n{ligne.get('designation','')}")
            c2.markdown(f"**{ligne.get('quantite','')}** {ligne.get('unite','')}")
            c3.markdown(fmt(ligne.get("prix_unitaire")))
            c4.markdown(f"**{fmt(ligne.get('total_ligne'))}**")
            if ligne.get("date_livraison_ligne"):
                st.caption(f"📅 Livraison ligne : {ligne['date_livraison_ligne']}")
            st.markdown("<hr style='margin:6px 0; border-color:#e0dfd8'>", unsafe_allow_html=True)

        t1, _, t3 = st.columns([3, 1, 2])
        t1.markdown("**Total HT**"); t3.markdown(f"**{fmt(d.get('total_ht'))}**")
        if d.get("tva_pct"):
            t1b, _, t3b = st.columns([3, 1, 2])
            t1b.markdown(f"TVA {d.get('tva_pct',20)} %")
            t3b.markdown(fmt((d.get("total_ht") or 0) * d.get("tva_pct", 20) / 100))
        if d.get("total_ttc"):
            t1c, _, t3c = st.columns([3, 1, 2])
            t1c.markdown("**Total TTC**"); t3c.markdown(f"**{fmt(d.get('total_ttc'))}**")


# ─── Interface principale ──────────────────────────────────────────────────────
st.title("📋 Extraction de bons de commande")
st.caption("Uploadez un BC (PDF ou image) — l'IA extrait automatiquement toutes les données utiles pour la saisie SAP.")

if not api_key:
    st.error("Clé API non configurée. Ajoutez ANTHROPIC_API_KEY dans les secrets Streamlit Cloud.")
    st.stop()

fichier = st.file_uploader("Déposez votre bon de commande", type=["pdf", "png", "jpg", "jpeg"],
                           help="PDF ou image (PNG, JPG). Taille max recommandée : 10 Mo.")

if fichier:
    col_btn, col_info = st.columns([1, 3])
    with col_btn:
        lancer = st.button("🔍 Analyser le BC", type="primary", use_container_width=True)
    with col_info:
        st.caption(f"Fichier : **{fichier.name}** · {fichier.size / 1024:.0f} Ko")

    if lancer:
        with st.spinner("Analyse en cours — Claude lit votre bon de commande..."):
            try:
                donnees = extraire_bc(fichier.read(), fichier.name, api_key)
                st.session_state["donnees_bc"] = donnees
                st.session_state["nom_fichier"] = fichier.name
                st.success("Extraction réussie !")
            except json.JSONDecodeError as e:
                st.error(f"Erreur de parsing JSON : {e}")
            except anthropic.AuthenticationError:
                st.error("Clé API invalide. Vérifiez votre clé Anthropic.")
            except Exception as e:
                st.error(f"Erreur : {e}")

if "donnees_bc" in st.session_state:
    donnees = st.session_state["donnees_bc"]
    st.markdown("---")
    afficher_resultats(donnees)
    st.markdown("---")

    col_ex, col_json, col_reset = st.columns([2, 2, 1])
    with col_ex:
        excel_bytes = generer_excel(donnees)
        if excel_bytes:
            nom_base = Path(st.session_state.get("nom_fichier", "bc")).stem
            st.download_button(label="📥 Télécharger Excel SAP", data=excel_bytes,
                               file_name=f"{nom_base}_SAP.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               type="primary", use_container_width=True)
    with col_json:
        with st.expander("Voir le JSON brut"):
            st.json(donnees)
    with col_reset:
        if st.button("🔄 Nouveau BC", use_container_width=True):
            del st.session_state["donnees_bc"]
            st.rerun()
